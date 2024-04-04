import openpyxl 
from openpyxl import Workbook
from openpyxl import load_workbook
import pandas as pd
from pandas import DataFrame
from openpyxl.utils.dataframe import dataframe_to_rows
import sys
import os

# Filter/Move 'Do Not Mail' contacts to 'Do Not Mail' Sheet
# do_not_mail_df = df[df["City"].str.contains("Los Angeles")]

'''
Take propstream data and get it ready for skiptracing
    Steps: 
    - Take out all property leads that have the same firstname/lastname (duplicate owners)
    - Take out leads with 'Do Not Mail' = No 
    - Format the propstream data according to LeadSherpa skiptrace template
        - First Name
        - Last name
        - Mail Address, City, State, Zip
        - Property Address, City, State, Zip
    - Delete all columns that are unnecessary:
        - APN Col
'''

# Adding Command Line Functionality
'''
Command to run program is 'python3 skiptrace_formatter.py input_filename output_filename'
'''
file1 = os.path.basename(sys.argv[1])
file2 = os.path.basename(sys.argv[2])

# Load existing workbook and assign sheet to variable
wb1 = load_workbook(file1)
ws = wb1["Sheet1"]

# Creating Output Workbook
wb2 = Workbook()
ws1 = wb2.active
ws1.title = "Sorted Data"

# Creating Dataframe to modify data with pandas library
df = DataFrame(ws.values, columns=['Owner 1 First Name', 'Owner 1 Last Name', 'Address', 'Unit #', 'City', 'State', 'Zip', 'Mailing Care of Name', 'Mailing Address', 'Mailing Unit #', 'Mailing City', 'Mailing State', 'Mailing Zip', 'APN', 'extra'])

# Renaming Certain Columns - FOR DATAFRAME ONLY
df.rename(columns={'Owner 1 First Name': 'FirstName', 'Owner 1 Last Name': 'LastName', 'Unit #': 'Unit', 'Mailing Unit #': 'Mailing Unit'}, inplace=True)

# Deleting Unnecessary Columns
df.drop(columns=["APN", "Mailing Care of Name", 'extra', 'Unit', 'Mailing Unit'], inplace=True)

# Repositioning Columns
df = df.reindex(columns=['FirstName', 'LastName', 'Mailing Address', 'Mailing City', 'Mailing State', 'Mailing Zip', 'Address', 'City', 'State', 'Zip'])

# Removing Rows based on Duplicate Owners
df.drop_duplicates(subset=["FirstName", "LastName"], keep='first', inplace=True)

# Changing dataframe back to workbook
for r in dataframe_to_rows(df, index=False, header=False):
    ws1.append(r)

# Resizing Workbook Cells
ws1.column_dimensions['A'].width = 20
ws1.column_dimensions['B'].width = 20
ws1.column_dimensions['C'].width = 25
ws1.column_dimensions['H'].width = 25

ws1.column_dimensions['D'].width = 12
ws1.column_dimensions['E'].width = 12
ws1.column_dimensions['F'].width = 12
ws1.column_dimensions['G'].width = 12

# Changing Workbook Header Cell Contents
ws1['A1'] = "FirstName"
ws1['B1'] = "LastName"

# final save of new workbook
wb2.save(sys.argv[2])