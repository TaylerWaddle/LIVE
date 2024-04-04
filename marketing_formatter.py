import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import pandas as pd
from pandas import DataFrame
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
import sys
import os

'''
Take skiptraced data and get it ready for Marketing List
    - Format into desired marketing List format(in this order): (delete columns not on this list)
        - First Name
        - Last Name
        - Property Address, City, State, Zip
        - Phone1, Phone Type
        - Phone2, Phone Type
        - Phone3, Phone Type
        - Add Call Status Col
        - Mailing Address, City, State, Zip
        - Add Mail Status Col
        - Vacant
        - Email1, Email2
        - Golden Address, City, State, Zip
        - Age
        - Is Deceased
        - Bankruptcy
        - Foreclosure
        - Lien
        - Judgement
        - Quitclaim
3) Automatically append Marketing List File with newly formatted leads
'''
# Adding Command Line Functionality
'''
Command to run program is 'python3 marketing_formatter.py input_filename output_filename'
'''
file1 = os.path.basename(sys.argv[1])
file2 = os.path.basename(sys.argv[2])
desired_sheet = sys.argv[3]

# Load existing workbook and assign sheet to variable
wb1 = load_workbook(file1)
ws = wb1["Sheet1"]

# Creating Output Workbook
wb2 = load_workbook(file2)
ws1 = wb2[desired_sheet]

# Creating Dataframe to modify data with pandas library
df = DataFrame(ws.values, columns=['First Name', 'Last Name', 
                                   'Mail Address', 'Mail City', 'Mail State', 'Mail Zip', 
                                   'Property Address', 'Property City', 'Property State', 'Property Zip', 
                                   'Custom 1', 'Custom 2', 'Custom 3', 'Custom 4', 'Custom 5', 'Custom 6',
                                   'Validated Mail Address', 'Validated Mail City', 'Validated Mail State', 'Validated Mail Zip', 
                                   'Vacant', 
                                   'Alternate Full Name', 'Alternate First Name', 'Alternate Last Name',
                                   'Phone1', 'Phone 1 Type', 'Phone 1 Last Seen', 'Phone2', 'Phone 2 Type', 'Phone 2 Last Seen', 'Phone3', 'Phone 3 Type', 'Phone 3 Last Seen',
                                   'Email1', 'Email1 Last Seen', 'Email2', 'Email2 Last Seen',
                                   'IP Address', 'IP Last Seen',
                                   'Golden Address', 'Golden City', 'Golden State', 'Golden Zip', 'Golden Address Last Seen',
                                   'Age', 'Is Deceased', 'Bankruptcy', 'Foreclosure', 'Lien', 'Judgement', 'Quitclaim', 'Available Equity',
                                   'Relative 1 First Name', 'Relative 1 Last Name', 'Relative 1 Phone1', 'Relative 1 Phone2', 'Relative 1 Phone3',
                                   'Relative 2 First Name', 'Relative 2 Last Name', 'Relative 2 Phone1', 'Relative 2 Phone2', 'Relative 2 Phone3',
                                   'Litigator', 'Has Hit', 'Record Status'])


# Deleting Unnecessary Columns
df.drop(columns=['Mail Address', 'Mail City', 'Mail State', 'Mail Zip', 
                 'Custom 1', 'Custom 2', 'Custom 3', 'Custom 4', 'Custom 5', 'Custom 6',
                 'Alternate Full Name', 'Alternate First Name', 'Alternate Last Name',
                 'Phone 1 Last Seen', 'Phone 2 Last Seen', 'Phone 3 Last Seen',
                 'Email1 Last Seen', 'Email2 Last Seen',
                 'IP Address', 'IP Last Seen',
                 'Golden Address Last Seen',
                 'Relative 1 First Name', 'Relative 1 Last Name', 'Relative 1 Phone1', 'Relative 1 Phone2', 'Relative 1 Phone3',
                 'Relative 2 First Name', 'Relative 2 Last Name', 'Relative 2 Phone1', 'Relative 2 Phone2', 'Relative 2 Phone3',
                 'Litigator', 'Has Hit', 'Record Status'], inplace=True)
df.drop([0], inplace=True)


# Renaming Certain Columns - FOR DATAFRAME ONLY
df.rename(columns={'Validated Mail Address': 'Mail Address', 'Validated Mail City': 'Mail City', 'Validated Mail State': 'Mail State', 'Validated Mail Zip': 'Mail Zip'}, inplace=True)


# Adding Missing Columns
call_status = []
mail_status = []
df['Call Status'] = pd.Series(call_status)
df['Mail Status'] = pd.Series(mail_status)


# Repositioning Columns
df = df.reindex(columns=['First Name', 'Last Name', 
                         'Property Address', 'Property City', 'Property State', 'Property Zip', 
                         'Phone1', 'Phone 1 Type', 'Phone2', 'Phone 2 Type', 'Phone3', 'Phone 3 Type',
                         'Call Status',
                         'Mail Address', 'Mail City', 'Mail State', 'Mail Zip',
                         'Mail Status',
                         'Vacant',
                         'Email1', 'Email2',
                         'Golden Address', 'Golden City', 'Golden State', 'Golden Zip',
                         'Age', 'Is Deceased',
                         'Bankruptcy', 'Foreclosure', 'Lien', 'Judgement', 'Quitclaim', 'Available Equity'])


# Changing dataframe back to workbook and appending to Marketing List
for row in dataframe_to_rows(df, index=False, header=False):
    ws1.append(row)

# final save of new workbook
wb2.save(sys.argv[2])




